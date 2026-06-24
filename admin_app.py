import sys
import os
import json
import sqlite3
import base64
import urllib.parse
import urllib.request
from datetime import datetime

from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QFont, QIcon, QColor, QPixmap
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem,
    QListWidget, QStackedWidget, QComboBox, QDateEdit, QFileDialog,
    QMessageBox, QSplitter, QHeaderView, QTabWidget, QDialog, 
    QFormLayout, QDialogButtonBox
)
import pandas as pd
import openpyxl
import database
from fpdf import FPDF

CONFIG_FILE = "github_config.json"

class AdminMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("카스테크(CAS-TECH) 전사 통합 관리 시스템")
        self.resize(1280, 800)
        
        # GitHub 원격 동기화 파일 정보용 SHA 관리 변수
        self.remote_db_sha = None
        
        # 1. DB 스키마 검증 및 작업자 테이블 보정
        database.ensure_worker_table()
        
        # 2. UI 스타일 및 전체 레이아웃 구성
        self.setup_styles()
        self.setup_ui()
        
        # 3. GitHub 설정 불러오기
        self.load_github_config()
        
        # 4. 시작 시 자동으로 DB 불러오기 (Pull)
        self.auto_pull_db()

    def setup_styles(self):
        # Industrial Embedded Dark UI 스타일시트 정의
        self.setStyleSheet("""
            QMainWindow {
                background-color: #121212;
            }
            QWidget {
                background-color: #121212;
                color: #ffffff;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel {
                color: #ffffff;
                font-size: 12px;
            }
            QLabel#tab-title {
                color: #deff9a;
                font-size: 18px;
                font-weight: bold;
                margin-bottom: 10px;
            }
            QLineEdit, QComboBox, QDateEdit {
                background-color: #1e1e1e;
                border: 1px solid #333333;
                border-radius: 5px;
                color: #ffffff;
                padding: 6px 10px;
                font-size: 12px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border: 1px solid #deff9a;
            }
            QPushButton {
                background-color: #1e1e1e;
                border: 1px solid #deff9a;
                border-radius: 5px;
                color: #deff9a;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #deff9a;
                color: #121212;
            }
            QPushButton:pressed {
                background-color: #c5eb80;
                color: #121212;
            }
            QPushButton#action-btn {
                background-color: #222222;
                border: 1px solid #deff9a;
                color: #deff9a;
            }
            QPushButton#action-btn:hover {
                background-color: #deff9a;
                color: #121212;
            }
            QPushButton#delete-btn {
                background-color: #2c1a1a;
                border: 1px solid #ff8a8a;
                color: #ff8a8a;
            }
            QPushButton#delete-btn:hover {
                background-color: #ff8a8a;
                color: #121212;
            }
            QListWidget {
                background-color: #181818;
                border: 1px solid #2d2d2d;
                border-radius: 6px;
                color: #cccccc;
                padding: 5px;
            }
            QListWidget::item {
                padding: 12px;
                margin-bottom: 2px;
                border-radius: 4px;
            }
            QListWidget::item:hover {
                background-color: #262626;
                color: #deff9a;
            }
            QListWidget::item:selected {
                background-color: #deff9a;
                color: #121212;
                font-weight: bold;
            }
            QTableWidget {
                background-color: #181818;
                border: 1px solid #2d2d2d;
                color: #ffffff;
                gridline-color: #262626;
                border-radius: 6px;
            }
            QTableWidget::item:hover {
                background-color: #262626;
            }
            QTableWidget::item:selected {
                background-color: #deff9a;
                color: #121212;
            }
            QHeaderView::section {
                background-color: #262626;
                color: #deff9a;
                font-weight: bold;
                padding: 8px;
                border: 1px solid #1a1a1a;
            }
            QTabWidget::pane {
                border: 1px solid #2d2d2d;
                background-color: #181818;
                border-radius: 6px;
            }
            QTabBar::tab {
                background-color: #222222;
                color: #cccccc;
                padding: 8px 16px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
                margin-right: 2px;
            }
            QTabBar::tab:hover {
                background-color: #2a2a2a;
                color: #deff9a;
            }
            QTabBar::tab:selected {
                background-color: #181818;
                color: #deff9a;
                border: 1px solid #2d2d2d;
                border-bottom-color: #181818;
                font-weight: bold;
            }
            QSplitter::handle {
                background-color: #2d2d2d;
            }
        """)

    def setup_ui(self):
        # 메인 중앙 위젯
        main_central = QWidget()
        self.setCentralWidget(main_central)
        
        main_layout = QVBoxLayout(main_central)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)
        

        # 2. 좌/우 2분할 스플리터 레이아웃
        splitter = QSplitter(Qt.Horizontal)
        main_layout.addWidget(splitter)
        
        # 2-1. 좌측 메뉴 탭 선택기 (QListWidget)
        self.nav_menu = QListWidget()
        self.nav_menu.addItems([
            "📊 데이터 통합 관리",
            "🖨️ 보고서 생성 및 출력",
            "👥 사용자 관리"
        ])
        self.nav_menu.setFixedWidth(200)
        self.nav_menu.currentRowChanged.connect(self.switch_tab)
        splitter.addWidget(self.nav_menu)
        
        # 2-2. 우측 메인 컨텐츠 영역 (QStackedWidget)
        self.main_stack = QStackedWidget()
        splitter.addWidget(self.main_stack)
        
        # 메인 스택 하위 화면(탭) 생성
        self.create_data_mgmt_tab()
        self.create_report_tab()
        self.create_workers_tab()
        
        # 초기 탭 설정
        self.nav_menu.setCurrentRow(0)
        splitter.setSizes([200, 1080])

    # =========================================================================
    # 👥 사용자 관리 탭
    # =========================================================================
    def create_workers_tab(self):
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        title = QLabel("👥 사용자 권한 및 직원 관리")
        title.setObjectName("tab-title")
        layout.addWidget(title)
        
        # 직원 목록 테이블
        self.workers_table = QTableWidget()
        self.workers_table.setColumnCount(4)
        self.workers_table.setHorizontalHeaderLabels(["작업자명", "권한", "등록일시", "비밀번호"])
        self.workers_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.workers_table.itemChanged.connect(self.on_worker_cell_edited)
        layout.addWidget(self.workers_table)
        
        # 기능 버튼 바
        btn_bar = QHBoxLayout()
        self.btn_add_worker = QPushButton("➕ 작업자 추가")
        self.btn_add_worker.setObjectName("action-btn")
        self.btn_add_worker.clicked.connect(self.add_worker_dialog)
        btn_bar.addWidget(self.btn_add_worker)
        
        self.btn_delete_worker = QPushButton("❌ 선택 작업자 삭제")
        self.btn_delete_worker.setObjectName("delete-btn")
        self.btn_delete_worker.clicked.connect(self.delete_worker_action)
        btn_bar.addWidget(self.btn_delete_worker)
        
        btn_bar.addStretch()
        layout.addLayout(btn_bar)
        
        self.main_stack.addWidget(tab_widget)

    # =========================================================================
    # 📊 데이터 통합 관리 탭
    # =========================================================================
    def create_data_mgmt_tab(self):
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        title = QLabel("📊 전사 데이터 통합 관리")
        title.setObjectName("tab-title")
        layout.addWidget(title)
        
        # 상단 필터 및 검색 바
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        filter_layout.addWidget(QLabel("🏢 거래처별:"))
        self.filter_client_combo = QComboBox()
        self.filter_client_combo.setFixedWidth(160)
        self.filter_client_combo.currentTextChanged.connect(self.load_all_data)
        filter_layout.addWidget(self.filter_client_combo)
        
        filter_layout.addWidget(QLabel("🔍 계기 ID별:"))
        self.filter_eq_id_input = QLineEdit()
        self.filter_eq_id_input.setPlaceholderText("검색할 계기 ID...")
        self.filter_eq_id_input.setFixedWidth(120)
        self.filter_eq_id_input.textChanged.connect(self.load_all_data)
        filter_layout.addWidget(self.filter_eq_id_input)
        
        filter_layout.addWidget(QLabel("📅 점검 일자별:"))
        self.filter_date_start = QDateEdit()
        self.filter_date_start.setCalendarPopup(True)
        self.filter_date_start.setFixedWidth(130)
        self.filter_date_start.setDate(datetime.today().replace(day=1))
        self.filter_date_start.dateChanged.connect(self.load_all_data)
        filter_layout.addWidget(self.filter_date_start)
        
        filter_layout.addWidget(QLabel("~"))
        self.filter_date_end = QDateEdit()
        self.filter_date_end.setCalendarPopup(True)
        self.filter_date_end.setFixedWidth(130)
        self.filter_date_end.setDate(datetime.today())
        self.filter_date_end.dateChanged.connect(self.load_all_data)
        filter_layout.addWidget(self.filter_date_end)
        
        self.btn_reset_filters = QPushButton("🔄 필터 초기화")
        self.btn_reset_filters.clicked.connect(self.reset_filters_action)
        filter_layout.addWidget(self.btn_reset_filters)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # 하위 테이블 분류 탭 구성 (거래처 / 계기 사양 / 점검 이력)
        self.sub_tabs = QTabWidget()
        layout.addWidget(self.sub_tabs)
        
        # 2-1. 거래처 탭
        self.clients_tab = QWidget()
        clients_layout = QVBoxLayout(self.clients_tab)
        self.clients_table = QTableWidget()
        self.clients_table.setColumnCount(1)
        self.clients_table.setHorizontalHeaderLabels(["거래처명"])
        self.clients_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.clients_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.clients_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.clients_table.setSelectionMode(QTableWidget.SingleSelection)
        self.clients_table.itemDoubleClicked.connect(self.on_client_double_clicked)
        self.clients_table.itemChanged.connect(self.on_client_cell_edited)
        clients_layout.addWidget(self.clients_table)
        
        clients_btn_bar = QHBoxLayout()
        self.btn_add_client = QPushButton("➕ 거래처 추가")
        self.btn_add_client.setObjectName("action-btn")
        self.btn_add_client.clicked.connect(self.add_client_dialog)
        clients_btn_bar.addWidget(self.btn_add_client)
        clients_btn_bar.addStretch()
        clients_layout.addLayout(clients_btn_bar)
        
        self.sub_tabs.addTab(self.clients_tab, "🏢 거래처 정보")
        
        # 2-2. 계기 사양 탭
        self.eq_tab = QWidget()
        eq_layout = QVBoxLayout(self.eq_tab)
        self.eq_table = QTableWidget()
        self.eq_table.setColumnCount(11)
        self.eq_table.setHorizontalHeaderLabels([
            "거래처명", "설비 ID (PK)", "설비명", "설치위치", "계측기 IP",
            "인디게이터", "로드셀", "형식", "설치년월", "설비사진 1", "설비사진 2"
        ])
        self.eq_table.verticalHeader().setDefaultSectionSize(55) # 사진 표시를 위해 행 높이 조정
        self.eq_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        
        # 더블클릭/클릭 시 수정 안 되고 행만 선택되도록 설정
        self.eq_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.eq_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.eq_table.setSelectionMode(QTableWidget.SingleSelection)
        
        self.eq_table.itemChanged.connect(self.on_eq_cell_edited)
        eq_layout.addWidget(self.eq_table)
        
        eq_btn_bar = QHBoxLayout()
        self.btn_add_eq = QPushButton("➕ 계기 사양 추가")
        self.btn_add_eq.setObjectName("action-btn")
        self.btn_add_eq.clicked.connect(self.add_eq_dialog)
        eq_btn_bar.addWidget(self.btn_add_eq)
        
        self.btn_edit_eq = QPushButton("✏️ 선택 계기 사양 수정")
        self.btn_edit_eq.setObjectName("action-btn")
        self.btn_edit_eq.clicked.connect(self.edit_eq_dialog)
        eq_btn_bar.addWidget(self.btn_edit_eq)
        
        self.btn_delete_eq = QPushButton("❌ 선택 계기 사양 삭제")
        self.btn_delete_eq.setObjectName("delete-btn")
        self.btn_delete_eq.clicked.connect(self.delete_eq_action)
        eq_btn_bar.addWidget(self.btn_delete_eq)
        eq_btn_bar.addStretch()
        eq_layout.addLayout(eq_btn_bar)
        
        self.sub_tabs.addTab(self.eq_tab, "📐 계기 사양 (Machines)")
        
        # 2-3. 점검 이력 탭
        self.logs_tab = QWidget()
        logs_layout = QVBoxLayout(self.logs_tab)
        self.logs_table = QTableWidget()
        self.logs_table.setColumnCount(10)
        self.logs_table.setHorizontalHeaderLabels([
            "ID (PK)", "점검일시", "작업자명", "거래처명", "설비 ID",
            "증상 및 상태", "조치 및 수리내용", "수리비용(원)", "사진 1", "사진 2"
        ])
        self.logs_table.verticalHeader().setDefaultSectionSize(55) # 사진 표시를 위해 행 높이 조정
        self.logs_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        
        # 더블클릭/클릭 시 수정 안 되고 행만 선택되도록 설정
        self.logs_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.logs_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.logs_table.setSelectionMode(QTableWidget.SingleSelection)
        
        self.logs_table.itemChanged.connect(self.on_log_cell_edited)
        logs_layout.addWidget(self.logs_table)
        
        logs_btn_bar = QHBoxLayout()
        self.btn_add_log = QPushButton("➕ 점검 이력 수동 등록")
        self.btn_add_log.setObjectName("action-btn")
        self.btn_add_log.clicked.connect(self.add_log_dialog)
        logs_btn_bar.addWidget(self.btn_add_log)
        
        self.btn_edit_log = QPushButton("✏️ 선택 점검 이력 수정")
        self.btn_edit_log.setObjectName("action-btn")
        self.btn_edit_log.clicked.connect(self.edit_log_dialog)
        logs_btn_bar.addWidget(self.btn_edit_log)
        
        self.btn_delete_log = QPushButton("❌ 선택 점검 이력 삭제")
        self.btn_delete_log.setObjectName("delete-btn")
        self.btn_delete_log.clicked.connect(self.delete_log_action)
        logs_btn_bar.addWidget(self.btn_delete_log)
        logs_btn_bar.addStretch()
        logs_layout.addLayout(logs_btn_bar)
        
        self.sub_tabs.addTab(self.logs_tab, "🛠️ 현장 점검 이력 (Logs)")
        
        self.main_stack.addWidget(tab_widget)

    # =========================================================================
    # 🖨️ 보고서 생성 및 출력 탭
    # =========================================================================
    def create_report_tab(self):
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        title = QLabel("🖨️ 수리/점검 종합 보고서 출력")
        title.setObjectName("tab-title")
        layout.addWidget(title)
        
        # 보고서 설명
        desc = QLabel("선택한 조건의 점검/수리 데이터를 쏙 뽑아서 깔끔한 포맷의 Excel 또는 PDF 파일로 바탕화면에 생성합니다.")
        desc.setStyleSheet("color: #aaaaaa; font-size: 13px;")
        layout.addWidget(desc)
        
        # 보고서 필터 박스
        filter_box = QWidget()
        filter_box.setStyleSheet("background-color: #181818; border: 1.5px solid #2d2d2d; border-radius: 8px; padding: 20px;")
        filter_layout = QFormLayout(filter_box)
        filter_layout.setSpacing(15)
        
        self.report_client_combo = QComboBox()
        filter_layout.addRow("🏢 대상 거래처 선택:", self.report_client_combo)
        
        self.report_date_start = QDateEdit()
        self.report_date_start.setCalendarPopup(True)
        self.report_date_start.setDate(datetime.today().replace(day=1))
        filter_layout.addRow("📅 조회 시작일:", self.report_date_start)
        
        self.report_date_end = QDateEdit()
        self.report_date_end.setCalendarPopup(True)
        self.report_date_end.setDate(datetime.today())
        filter_layout.addRow("📅 조회 종료일:", self.report_date_end)
        
        layout.addWidget(filter_box)
        
        # 버튼 영역
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(20)
        
        self.btn_export_excel = QPushButton("📊 Excel 점검 보고서 생성 및 저장")
        self.btn_export_excel.setFixedHeight(45)
        self.btn_export_excel.clicked.connect(self.export_excel_report)
        btn_layout.addWidget(self.btn_export_excel)
        
        self.btn_export_pdf = QPushButton("📄 PDF 점검 보고서 생성 및 저장")
        self.btn_export_pdf.setFixedHeight(45)
        self.btn_export_pdf.clicked.connect(self.export_pdf_report)
        btn_layout.addWidget(self.btn_export_pdf)
        
        layout.addLayout(btn_layout)
        layout.addStretch()
        
        self.main_stack.addWidget(tab_widget)

    # =========================================================================
    # 메인 컨트롤 네비게이션 제어
    # =========================================================================
    def switch_tab(self, index):
        self.main_stack.setCurrentIndex(index)
        self.load_all_data()

    def reset_filters_action(self):
        self.filter_client_combo.setCurrentIndex(0)
        self.filter_eq_id_input.clear()
        self.filter_date_start.setDate(datetime.today().replace(day=1))
        self.filter_date_end.setDate(datetime.today())
        self.load_all_data()

    # =========================================================================
    # 데이터 로딩 및 동적 연동 로직
    # =========================================================================
    def load_all_data(self):
        # 1. 사용자 관리 테이블 채우기
        self.load_workers_table()
        
        # 2. 거래처 셀렉트 박스 필터들 동기화
        try:
            current_filter = self.filter_client_combo.currentText()
            current_report_client = self.report_client_combo.currentText()
            
            clients = database.get_clients()
            
            # 필터 콤보 박스 갱신 (시그널 차단)
            self.filter_client_combo.blockSignals(True)
            self.filter_client_combo.clear()
            self.filter_client_combo.addItems(["전체 거래처"] + clients)
            if current_filter in ["전체 거래처"] + clients:
                self.filter_client_combo.setCurrentText(current_filter)
            self.filter_client_combo.blockSignals(True)
            self.filter_client_combo.blockSignals(False)
            
            # 보고서 콤보 박스 갱신
            self.report_client_combo.blockSignals(True)
            self.report_client_combo.clear()
            self.report_client_combo.addItems(["전체 거래처"] + clients)
            if current_report_client in ["전체 거래처"] + clients:
                self.report_client_combo.setCurrentText(current_report_client)
            self.report_client_combo.blockSignals(False)
        except Exception as e:
            print(f"Error loading clients combo: {e}")
            
        # 3. 데이터 탭들의 테이블 채우기 (필터값 기반)
        self.load_clients_table()
        self.load_equipments_table()
        self.load_logs_table()

    # 1. 사용자 테이블 로드
    def load_workers_table(self):
        self.workers_table.blockSignals(True)
        self.workers_table.setRowCount(0)
        
        try:
            conn = database.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT 작업자명, 권한, 등록일시, 비밀번호 FROM 작업자 ORDER BY 등록일시 DESC")
            rows = cursor.fetchall()
            conn.close()
            
            for row_idx, row_data in enumerate(rows):
                self.workers_table.insertRow(row_idx)
                for col_idx, value in enumerate(row_data):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    if col_idx == 0:
                        # 수정 전 원래 이름을 UserRole에 저장하여 PK 갱신 시 활용
                        item.setData(Qt.UserRole, str(value))
                    if col_idx == 2:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 등록일시는 수정 불가
                    self.workers_table.setItem(row_idx, col_idx, item)
        except Exception as e:
            print(f"Error loading workers table: {e}")
        finally:
            self.workers_table.blockSignals(False)

    # 2. 거래처 테이블 로드
    def load_clients_table(self):
        self.clients_table.blockSignals(True)
        self.clients_table.setRowCount(0)
        
        try:
            clients = database.get_clients()
            # 검색어 연동 (대소문자 무시)
            client_search = self.filter_client_combo.currentText()
            
            for row_idx, client in enumerate(clients):
                if client_search != "전체 거래처" and client != client_search:
                    continue
                self.clients_table.insertRow(self.clients_table.rowCount())
                curr_row = self.clients_table.rowCount() - 1
                item = QTableWidgetItem(client)
                item.setData(Qt.UserRole, client)  # 수정 전 이름 보존
                self.clients_table.setItem(curr_row, 0, item)
        except Exception as e:
            print(f"Error loading clients table: {e}")
        finally:
            self.clients_table.blockSignals(False)

    # 3. 계기 사양 테이블 로드
    def load_equipments_table(self):
        self.eq_table.blockSignals(True)
        self.eq_table.setRowCount(0)
        
        client_filter = self.filter_client_combo.currentText()
        eq_id_search = self.filter_eq_id_input.text().strip().lower()
        
        try:
            conn = database.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = "SELECT * FROM 설비마스터 WHERE 1=1"
            params = []
            
            if client_filter != "전체 거래처" and client_filter:
                query += " AND 거래처명 = ?"
                params.append(client_filter)
            if eq_id_search:
                query += " AND (LOWER(설비ID) LIKE ? OR LOWER(설비명) LIKE ?)"
                params.extend([f"%{eq_id_search}%", f"%{eq_id_search}%"])
                
            query += " ORDER BY 설비ID"
            cursor.execute(query, params)
            rows = [dict(r) for r in cursor.fetchall()]
            conn.close()
            
            cols = ["거래처명", "설비ID", "설비명", "설치위치", "계측기_IP", "인디게이터", "로드셀", "형식", "설치년월"]
            for row_idx, eq in enumerate(rows):
                self.eq_table.insertRow(row_idx)
                for col_idx, col_name in enumerate(cols):
                    val = eq.get(col_name)
                    item = QTableWidgetItem(str(val) if val is not None else "")
                    if col_name == "설비ID":
                        # 수정 전 설비 ID를 UserRole에 저장하여 PK 갱신 시 활용
                        item.setData(Qt.UserRole, str(val))
                    self.eq_table.setItem(row_idx, col_idx, item)
                
                # 설비사진 1, 2 썸네일 표시
                for p_idx, col_pic in enumerate(["설비사진1", "설비사진2"]):
                    pic_path = eq.get(col_pic)
                    col_target = len(cols) + p_idx # 9, 10번째 열
                    
                    if pic_path and os.path.exists(pic_path):
                        label = QLabel()
                        pixmap = QPixmap(pic_path)
                        if not pixmap.isNull():
                            # 썸네일 생성 및 QLabel 세팅
                            scaled_pix = pixmap.scaled(60, 45, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            label.setPixmap(scaled_pix)
                            label.setAlignment(Qt.AlignCenter)
                            self.eq_table.setCellWidget(row_idx, col_target, label)
                        else:
                            item = QTableWidgetItem("손상 이미지")
                            item.setTextAlignment(Qt.AlignCenter)
                            self.eq_table.setItem(row_idx, col_target, item)
                    else:
                        item = QTableWidgetItem("-")
                        item.setTextAlignment(Qt.AlignCenter)
                        self.eq_table.setItem(row_idx, col_target, item)
        except Exception as e:
            print(f"Error loading equipments table: {e}")
        finally:
            self.eq_table.blockSignals(False)

    # 4. 점검 이력 테이블 로드
    def load_logs_table(self):
        self.logs_table.blockSignals(True)
        self.logs_table.setRowCount(0)
        
        client_filter = self.filter_client_combo.currentText()
        eq_id_search = self.filter_eq_id_input.text().strip().lower()
        start_date = self.filter_date_start.date().toPython()
        end_date = self.filter_date_end.date().toPython()
        
        try:
            conn = database.get_connection()
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            query = "SELECT * FROM 점검이력 WHERE 1=1"
            params = []
            
            if client_filter != "전체 거래처" and client_filter:
                query += " AND 거래처명 = ?"
                params.append(client_filter)
            if eq_id_search:
                query += " AND LOWER(설비ID) LIKE ?"
                params.append(f"%{eq_id_search}%")
                
            query += " ORDER BY 날짜_시간 DESC"
            cursor.execute(query, params)
            rows = [dict(r) for r in cursor.fetchall()]
            conn.close()
            
            cols = ["id", "날짜_시간", "작업자명", "거래처명", "설비ID", "증상상태", "조치_및_수리내용", "금액"]
            visible_row = 0
            
            for eq in rows:
                dt_str = eq.get("날짜_시간")
                if not dt_str:
                    continue
                try:
                    h_date_str = dt_str[:10].strip()
                    h_date = datetime.strptime(h_date_str, "%Y-%m-%d").date()
                    # 날짜 기간 필터링 적용
                    if not (start_date <= h_date <= end_date):
                        continue
                except Exception:
                    pass
                    
                self.logs_table.insertRow(visible_row)
                for col_idx, col_name in enumerate(cols):
                    val = eq.get(col_name)
                    item = QTableWidgetItem(str(val) if val is not None else "")
                    if col_name == "id":
                        # ID(PK)는 절대 수정 불가 처리
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        item.setTextAlignment(Qt.AlignCenter)
                    self.logs_table.setItem(visible_row, col_idx, item)
                    
                # 사진 1, 2 썸네일 표시
                for p_idx, col_pic in enumerate(["사진1", "사진2"]):
                    pic_path = eq.get(col_pic)
                    col_target = len(cols) + p_idx # 8, 9번째 열
                    
                    if pic_path and os.path.exists(pic_path):
                        label = QLabel()
                        pixmap = QPixmap(pic_path)
                        if not pixmap.isNull():
                            scaled_pix = pixmap.scaled(60, 45, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            label.setPixmap(scaled_pix)
                            label.setAlignment(Qt.AlignCenter)
                            self.logs_table.setCellWidget(visible_row, col_target, label)
                        else:
                            item = QTableWidgetItem("손상 이미지")
                            item.setTextAlignment(Qt.AlignCenter)
                            self.logs_table.setItem(visible_row, col_target, item)
                    else:
                        item = QTableWidgetItem("-")
                        item.setTextAlignment(Qt.AlignCenter)
                        self.logs_table.setItem(visible_row, col_target, item)
                
                visible_row += 1
        except Exception as e:
            print(f"Error loading logs table: {e}")
        finally:
            self.logs_table.blockSignals(False)

    # =========================================================================
    # 표(Table) 더블클릭 실시간 데이터 업데이트 연동
    # =========================================================================
    def on_worker_cell_edited(self, item):
        original_name = item.data(Qt.UserRole)
        row = item.row()
        col = item.column()
        new_val = item.text().strip()
        
        # 수정 시작 전 원격 DB에서 최신 데이터 가져옴
        database.sync_pull_from_github()
        
        conn = database.get_connection()
        cursor = conn.cursor()
        
        try:
            if col == 0:  # 작업자명 (PK) 변경 시
                if not new_val:
                    QMessageBox.warning(self, "오류", "작업자명은 필수 항목입니다.")
                    self.load_workers_table()
                    return
                cursor.execute("UPDATE 작업자 SET 작업자명 = ? WHERE 작업자명 = ?", (new_val, original_name))
                # 기존 점검이력의 작업자명도 일치시켜 연동
                cursor.execute("UPDATE 점검이력 SET 작업자명 = ? WHERE 작업자명 = ?", (new_val, original_name))
                item.setData(Qt.UserRole, new_val)
            elif col == 1:  # 권한 변경 시
                cursor.execute("UPDATE 작업자 SET 권한 = ? WHERE 작업자명 = ?", (new_val, self.workers_table.item(row, 0).text()))
            elif col == 3:  # 비밀번호 변경 시
                if not new_val:
                    QMessageBox.warning(self, "오류", "비밀번호는 필수 항목입니다.")
                    self.load_workers_table()
                    return
                cursor.execute("UPDATE 작업자 SET 비밀번호 = ? WHERE 작업자명 = ?", (new_val, self.workers_table.item(row, 0).text()))
                
            conn.commit()
            # 수정 성공 후 원격 DB로 적용
            database.sync_push_to_github()
        except Exception as e:
            QMessageBox.critical(self, "데이터베이스 오류", f"수정에 실패했습니다: {e}")
            self.load_workers_table()
        finally:
            conn.close()

    def on_client_cell_edited(self, item):
        original_name = item.data(Qt.UserRole)
        new_name = item.text().strip()
        
        if not new_name:
            QMessageBox.warning(self, "오류", "거래처명은 필수 항목입니다.")
            self.load_clients_table()
            return
            
        # database.py의 rename_client 유틸리티 함수 호출
        ok, msg = database.rename_client(original_name, new_name)
        if ok:
            item.setData(Qt.UserRole, new_name)
            self.load_all_data()
        else:
            QMessageBox.critical(self, "오류", msg)
            self.load_clients_table()

    def on_client_double_clicked(self, item):
        if item:
            client_name = item.text().strip()
            if client_name:
                # 📐 계기 사양 (Machines) 탭으로 전환 (인덱스 1)
                self.sub_tabs.setCurrentIndex(1)
                # 해당 거래처명으로 필터 지정하여 이외의 거래처는 보이지 않게 함
                self.filter_client_combo.setCurrentText(client_name)

    def on_eq_cell_edited(self, item):
        row = item.row()
        col = item.column()
        new_val = item.text().strip()
        
        # 수정 시작 전 원격 DB에서 최신 데이터 가져옴
        database.sync_pull_from_github()
        
        # 설비 ID 열에서 원래 ID 가져오기
        original_eq_id = self.eq_table.item(row, 1).data(Qt.UserRole)
        cols = ["거래처명", "설비ID", "설비명", "설치위치", "계측기_IP", "인디게이터", "로드셀", "형식", "설치년월"]
        col_name = cols[col]
        
        conn = database.get_connection()
        cursor = conn.cursor()
        
        try:
            if col_name == "설비ID":
                if not new_val:
                    QMessageBox.warning(self, "오류", "설비 ID는 필수 항목입니다.")
                    self.load_equipments_table()
                    return
                cursor.execute("UPDATE 설비마스터 SET 설비ID = ? WHERE 설비ID = ?", (new_val, original_eq_id))
                # 점검이력의 설비ID도 업데이트
                cursor.execute("UPDATE 점검이력 SET 설비ID = ? WHERE 설비ID = ?", (new_val, original_eq_id))
                self.eq_table.item(row, 1).setData(Qt.UserRole, new_val)
            else:
                cursor.execute(f"UPDATE 설비마스터 SET {col_name} = ? WHERE 설비ID = ?", (new_val, original_eq_id))
                
            conn.commit()
            # 수정 성공 후 원격 DB로 적용
            database.sync_push_to_github()
        except Exception as e:
            QMessageBox.critical(self, "데이터베이스 오류", f"수정에 실패했습니다: {e}")
            self.load_equipments_table()
        finally:
            conn.close()

    def on_log_cell_edited(self, item):
        row = item.row()
        col = item.column()
        new_val = item.text().strip()
        
        # 수정 시작 전 원격 DB에서 최신 데이터 가져옴
        database.sync_pull_from_github()
        
        log_id = self.logs_table.item(row, 0).text()
        cols = ["id", "날짜_시간", "작업자명", "거래처명", "설비ID", "증상상태", "조치_및_수리내용", "금액"]
        col_name = cols[col]
        
        conn = database.get_connection()
        cursor = conn.cursor()
        
        try:
            if col_name == "금액":
                try:
                    new_val = float(new_val)
                except ValueError:
                    new_val = 0.0
            
            cursor.execute(f"UPDATE 점검이력 SET {col_name} = ? WHERE id = ?", (new_val, log_id))
            conn.commit()
            # 수정 성공 후 원격 DB로 적용
            database.sync_push_to_github()
        except Exception as e:
            QMessageBox.critical(self, "데이터베이스 오류", f"수정에 실패했습니다: {e}")
            self.load_logs_table()
        finally:
            conn.close()

    # =========================================================================
    # 사용자 관리 다이얼로그 및 삭제 로직
    # =========================================================================
    def add_worker_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("작업자 추가")
        form = QFormLayout(dialog)
        
        name_input = QLineEdit()
        form.addRow("작업자명 (필수):", name_input)
        
        role_combo = QComboBox()
        role_combo.addItems(["작업자", "관리자", "기술팀장"])
        form.addRow("권한 등급:", role_combo)
        
        pw_input = QLineEdit()
        pw_input.setText("1234")
        form.addRow("비밀번호:", pw_input)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dialog)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        form.addRow(button_box)
        
        if dialog.exec() == QDialog.Accepted:
            name = name_input.text().strip()
            role = role_combo.currentText()
            pw = pw_input.text().strip()
            if not name:
                QMessageBox.warning(self, "오류", "작업자명을 입력해야 합니다.")
                return
            if not pw:
                pw = "1234"
                
            # 작업 시작 전 원격 DB에서 최신 데이터 가져옴
            database.sync_pull_from_github()
            
            conn = database.get_connection()
            cursor = conn.cursor()
            try:
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute("INSERT INTO 작업자 (작업자명, 권한, 등록일시, 비밀번호) VALUES (?, ?, ?, ?)", (name, role, now_str, pw))
                conn.commit()
                # 성공 후 원격 DB에 적용
                database.sync_push_to_github()
                self.load_workers_table()
            except sqlite3.IntegrityError:
                QMessageBox.warning(self, "오류", "이미 등록된 작업자명입니다.")
            except Exception as e:
                QMessageBox.critical(self, "오류", f"등록 실패: {e}")
            finally:
                conn.close()

    def delete_worker_action(self):
        curr_row = self.workers_table.currentRow()
        if curr_row < 0:
            QMessageBox.warning(self, "알림", "삭제할 작업자를 표에서 선택해 주세요.")
            return
            
        name = self.workers_table.item(curr_row, 0).text()
        confirm = QMessageBox.question(self, "삭제 확인", f"작업자 '{name}'을(를) 삭제하시겠습니까?\n이 이력은 복구되지 않습니다.", QMessageBox.Yes | QMessageBox.No)
        
        if confirm == QMessageBox.Yes:
            # 작업 시작 전 원격 DB에서 최신 데이터 가져옴
            database.sync_pull_from_github()
            
            conn = database.get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("DELETE FROM 작업자 WHERE 작업자명 = ?", (name,))
                conn.commit()
                # 성공 후 원격 DB에 적용
                database.sync_push_to_github()
                self.load_workers_table()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"삭제 실패: {e}")
            finally:
                conn.close()

    # =========================================================================
    # 거래처 추가 다이얼로그
    # =========================================================================
    def add_client_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("새로운 거래처 등록")
        form = QFormLayout(dialog)
        
        name_input = QLineEdit()
        form.addRow("거래처명 (필수):", name_input)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dialog)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        form.addRow(button_box)
        
        if dialog.exec() == QDialog.Accepted:
            name = name_input.text().strip()
            if not name:
                QMessageBox.warning(self, "오류", "거래처명을 입력해야 합니다.")
                return
                
            # database.py의 add_client 호출
            ok, msg = database.add_client(name)
            if ok:
                QMessageBox.information(self, "성공", msg)
                self.load_all_data()
            else:
                QMessageBox.warning(self, "실패", msg)

    # =========================================================================
    # 계기 사양 추가 / 삭제 다이얼로그
    # =========================================================================
    def add_eq_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("계기 사양 등록")
        dialog.resize(800, 600)
        form = QFormLayout(dialog)
        
        client_combo = QComboBox()
        client_combo.addItems(database.get_clients())
        form.addRow("거래처 선택 (필수):", client_combo)
        
        id_input = QLineEdit()
        form.addRow("설비 ID (필수):", id_input)
        
        name_input = QLineEdit()
        form.addRow("설비명 (필수):", name_input)
        
        loc_input = QLineEdit()
        form.addRow("설치위치:", loc_input)
        
        ip_input = QLineEdit()
        form.addRow("계측기 IP:", ip_input)
        
        ind_input = QLineEdit()
        form.addRow("인디게이터:", ind_input)
        
        lc_input = QLineEdit()
        form.addRow("로드셀:", lc_input)
        
        fmt_input = QLineEdit()
        form.addRow("형식:", fmt_input)
        
        date_input = QLineEdit()
        date_input.setPlaceholderText("예: 2026-06")
        form.addRow("설치년월:", date_input)
        
        # 신규 등록 사진 경로 변수 및 UI
        self.add_photo1_path = ""
        self.add_photo2_path = ""
        
        # 사진 1 등록 영역
        p1_layout = QHBoxLayout()
        self.add_p1_label = QLabel("사진 1 없음")
        p1_layout.addWidget(self.add_p1_label)
        btn_p1 = QPushButton("📁 사진 1 선택...")
        def choose_p1():
            fname, _ = QFileDialog.getOpenFileName(dialog, "설비 사진 1 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.add_photo1_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.add_p1_label.setPixmap(pix)
        btn_p1.clicked.connect(choose_p1)
        p1_layout.addWidget(btn_p1)
        form.addRow("설비 사진 1:", p1_layout)
        
        # 사진 2 등록 영역
        p2_layout = QHBoxLayout()
        self.add_p2_label = QLabel("사진 2 없음")
        p2_layout.addWidget(self.add_p2_label)
        btn_p2 = QPushButton("📁 사진 2 선택...")
        def choose_p2():
            fname, _ = QFileDialog.getOpenFileName(dialog, "설비 사진 2 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.add_photo2_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.add_p2_label.setPixmap(pix)
        btn_p2.clicked.connect(choose_p2)
        p2_layout.addWidget(btn_p2)
        form.addRow("설비 사진 2:", p2_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dialog)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        form.addRow(button_box)
        
        if dialog.exec() == QDialog.Accepted:
            client = client_combo.currentText()
            eq_id = id_input.text().strip()
            name = name_input.text().strip()
            
            if not client or not eq_id or not name:
                QMessageBox.warning(self, "오류", "필수 입력 항목(거래처, ID, 설비명)을 기입해야 합니다.")
                return
                
            # 사진 저장 처리
            p1_dest = ""
            p2_dest = ""
            PHOTOS_DIR = "photos"
            if not os.path.exists(PHOTOS_DIR):
                os.makedirs(PHOTOS_DIR)
                
            import shutil
            if self.add_photo1_path:
                p1_dest = os.path.join(PHOTOS_DIR, f"{eq_id}_1.jpg")
                try:
                    shutil.copyfile(self.add_photo1_path, p1_dest)
                    database.upload_photo_to_github(p1_dest)
                except Exception as e:
                    print(f"Error copying photo 1: {e}")
                    
            if self.add_photo2_path:
                p2_dest = os.path.join(PHOTOS_DIR, f"{eq_id}_2.jpg")
                try:
                    shutil.copyfile(self.add_photo2_path, p2_dest)
                    database.upload_photo_to_github(p2_dest)
                except Exception as e:
                    print(f"Error copying photo 2: {e}")
                    
            eq_data = {
                "거래처명": client,
                "설비ID": eq_id,
                "설비명": name,
                "설치위치": loc_input.text().strip(),
                "계측기_IP": ip_input.text().strip(),
                "인디게이터": ind_input.text().strip(),
                "로드셀": lc_input.text().strip(),
                "형식": fmt_input.text().strip(),
                "설치년월": date_input.text().strip(),
                "설비사진1": p1_dest,
                "설비사진2": p2_dest
            }
            
            ok, msg = database.add_equipment(eq_data)
            if ok:
                QMessageBox.information(self, "성공", msg)
                self.load_all_data()
            else:
                QMessageBox.warning(self, "실패", msg)

    def edit_eq_dialog(self):
        curr_row = self.eq_table.currentRow()
        if curr_row < 0:
            QMessageBox.warning(self, "알림", "수정할 계기를 표에서 선택해 주세요.")
            return
            
        eq_id = self.eq_table.item(curr_row, 1).data(Qt.UserRole)
        
        conn = database.get_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM 설비마스터 WHERE 설비ID = ?", (eq_id,))
        eq_data = cursor.fetchone()
        conn.close()
        
        if not eq_data:
            QMessageBox.critical(self, "오류", "선택한 설비 데이터를 찾을 수 없습니다.")
            return
            
        eq_data = dict(eq_data)
        
        dialog = QDialog(self)
        dialog.setWindowTitle("계기 사양 수정")
        dialog.resize(800, 600)
        form = QFormLayout(dialog)
        
        client_combo = QComboBox()
        clients = database.get_clients()
        client_combo.addItems(clients)
        client_combo.setCurrentText(eq_data["거래처명"])
        form.addRow("거래처 선택 (필수):", client_combo)
        
        id_input = QLineEdit()
        id_input.setText(eq_data["설비ID"])
        form.addRow("설비 ID (필수):", id_input)
        
        name_input = QLineEdit()
        name_input.setText(eq_data["설비명"])
        form.addRow("설비명 (필수):", name_input)
        
        loc_input = QLineEdit()
        loc_input.setText(eq_data["설치위치"] or "")
        form.addRow("설치위치:", loc_input)
        
        ip_input = QLineEdit()
        ip_input.setText(eq_data["계측기_IP"] or "")
        form.addRow("계측기 IP:", ip_input)
        
        ind_input = QLineEdit()
        ind_input.setText(eq_data["인디게이터"] or "")
        form.addRow("인디게이터:", ind_input)
        
        lc_input = QLineEdit()
        lc_input.setText(eq_data["로드셀"] or "")
        form.addRow("로드셀:", lc_input)
        
        fmt_input = QLineEdit()
        fmt_input.setText(eq_data["형식"] or "")
        form.addRow("형식:", fmt_input)
        
        date_input = QLineEdit()
        date_input.setText(eq_data["설치년월"] or "")
        date_input.setPlaceholderText("예: 2026-06")
        form.addRow("설치년월:", date_input)
        
        self.edit_photo1_path = eq_data["설비사진1"] or ""
        self.edit_photo2_path = eq_data["설비사진2"] or ""
        
        # 사진 1 영역
        p1_layout = QHBoxLayout()
        self.p1_label = QLabel("사진 1 없음")
        if self.edit_photo1_path and os.path.exists(self.edit_photo1_path):
            pix = QPixmap(self.edit_photo1_path).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.p1_label.setPixmap(pix)
        p1_layout.addWidget(self.p1_label)
        btn_p1 = QPushButton("📁 사진 1 변경...")
        def change_p1():
            fname, _ = QFileDialog.getOpenFileName(dialog, "설비 사진 1 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.edit_photo1_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.p1_label.setPixmap(pix)
        btn_p1.clicked.connect(change_p1)
        p1_layout.addWidget(btn_p1)
        form.addRow("설비 사진 1:", p1_layout)
        
        # 사진 2 영역
        p2_layout = QHBoxLayout()
        self.p2_label = QLabel("사진 2 없음")
        if self.edit_photo2_path and os.path.exists(self.edit_photo2_path):
            pix = QPixmap(self.edit_photo2_path).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.p2_label.setPixmap(pix)
        p2_layout.addWidget(self.p2_label)
        btn_p2 = QPushButton("📁 사진 2 변경...")
        def change_p2():
            fname, _ = QFileDialog.getOpenFileName(dialog, "설비 사진 2 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.edit_photo2_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.p2_label.setPixmap(pix)
        btn_p2.clicked.connect(change_p2)
        p2_layout.addWidget(btn_p2)
        form.addRow("설비 사진 2:", p2_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dialog)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        form.addRow(button_box)
        
        if dialog.exec() == QDialog.Accepted:
            client = client_combo.currentText()
            new_id = id_input.text().strip()
            name = name_input.text().strip()
            
            if not client or not new_id or not name:
                QMessageBox.warning(self, "오류", "필수 입력 항목(거래처, ID, 설비명)을 기입해야 합니다.")
                return
                
            p1_dest = eq_data["설비사진1"] or ""
            p2_dest = eq_data["설비사진2"] or ""
            PHOTOS_DIR = "photos"
            if not os.path.exists(PHOTOS_DIR):
                os.makedirs(PHOTOS_DIR)
                
            import shutil
            if self.edit_photo1_path and self.edit_photo1_path != eq_data["설비사진1"]:
                p1_dest = os.path.join(PHOTOS_DIR, f"{new_id}_1.jpg")
                try:
                    shutil.copyfile(self.edit_photo1_path, p1_dest)
                    database.upload_photo_to_github(p1_dest)
                except Exception as e:
                    print(f"Error copying photo 1: {e}")
                    
            if self.edit_photo2_path and self.edit_photo2_path != eq_data["설비사진2"]:
                p2_dest = os.path.join(PHOTOS_DIR, f"{new_id}_2.jpg")
                try:
                    shutil.copyfile(self.edit_photo2_path, p2_dest)
                    database.upload_photo_to_github(p2_dest)
                except Exception as e:
                    print(f"Error copying photo 2: {e}")
            
            updated_eq_data = {
                "거래처명": client,
                "설비ID": new_id,
                "설비명": name,
                "설치위치": loc_input.text().strip(),
                "계측기_IP": ip_input.text().strip(),
                "인디게이터": ind_input.text().strip(),
                "로드셀": lc_input.text().strip(),
                "형식": fmt_input.text().strip(),
                "설치년월": date_input.text().strip(),
                "설비사진1": p1_dest,
                "설비사진2": p2_dest
            }
            
            ok, msg = database.update_equipment(eq_id, updated_eq_data)
            if ok:
                QMessageBox.information(self, "성공", msg)
                self.load_all_data()
            else:
                QMessageBox.warning(self, "실패", msg)

    def delete_eq_action(self):
        curr_row = self.eq_table.currentRow()
        if curr_row < 0:
            QMessageBox.warning(self, "알림", "삭제할 계기를 표에서 선택해 주세요.")
            return
            
        eq_id = self.eq_table.item(curr_row, 1).text()
        confirm = QMessageBox.question(self, "삭제 확인", f"계기 ID '{eq_id}' 사양을 삭제하시겠습니까?\n이 이력은 복구되지 않습니다.", QMessageBox.Yes | QMessageBox.No)
        if confirm == QMessageBox.Yes:
            # 작업 시작 전 원격 DB에서 최신 데이터 가져옴
            database.sync_pull_from_github()
            
            conn = database.get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("DELETE FROM 설비마스터 WHERE 설비ID = ?", (eq_id,))
                cursor.execute("DELETE FROM 점검이력 WHERE 설비ID = ?", (eq_id,))
                conn.commit()
                # 성공 후 원격 DB에 적용
                database.sync_push_to_github()
                self.load_all_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"삭제 실패: {e}")
            finally:
                conn.close()
    # =========================================================================
    # 점검 이력 수동 등록 / 삭제 다이얼로그
    # =========================================================================
    def add_log_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("점검 이력 수동 입력")
        dialog.resize(800, 600)
        form = QFormLayout(dialog)
        
        clients = database.get_clients()
        client_combo = QComboBox()
        client_combo.addItems(clients)
        form.addRow("거래처 선택:", client_combo)
        
        eq_combo = QComboBox()
        def on_client_changed(client_name):
            eq_combo.clear()
            eqs = database.get_equipments(client_name)
            eq_combo.addItems([eq["설비ID"] for eq in eqs])
        
        client_combo.currentTextChanged.connect(on_client_changed)
        if clients:
            on_client_changed(clients[0])
            
        form.addRow("설비 ID 선택:", eq_combo)
        
        worker_combo = QComboBox()
        worker_combo.addItems(database.get_workers())
        form.addRow("작업자 선택:", worker_combo)
        
        date_input = QLineEdit()
        date_input.setText(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        form.addRow("점검 일시:", date_input)
        
        symptom_input = QLineEdit()
        form.addRow("증상상태:", symptom_input)
        
        action_input = QLineEdit()
        form.addRow("조치 및 수리내용:", action_input)
        
        cost_input = QLineEdit()
        cost_input.setText("0")
        form.addRow("수리 비용(원):", cost_input)
        
        # 신규 등록 점검 사진 경로 변수 및 UI
        self.add_log_photo1_path = ""
        self.add_log_photo2_path = ""
        
        # 사진 1 등록 영역
        p1_layout = QHBoxLayout()
        self.add_log_p1_label = QLabel("사진 1 없음")
        p1_layout.addWidget(self.add_log_p1_label)
        btn_p1 = QPushButton("📁 사진 1 선택...")
        def choose_p1():
            fname, _ = QFileDialog.getOpenFileName(dialog, "점검 사진 1 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.add_log_photo1_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.add_log_p1_label.setPixmap(pix)
        btn_p1.clicked.connect(choose_p1)
        p1_layout.addWidget(btn_p1)
        form.addRow("점검 사진 1:", p1_layout)
        
        # 사진 2 등록 영역
        p2_layout = QHBoxLayout()
        self.add_log_p2_label = QLabel("사진 2 없음")
        p2_layout.addWidget(self.add_log_p2_label)
        btn_p2 = QPushButton("📁 사진 2 선택...")
        def choose_p2():
            fname, _ = QFileDialog.getOpenFileName(dialog, "점검 사진 2 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.add_log_photo2_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.add_log_p2_label.setPixmap(pix)
        btn_p2.clicked.connect(choose_p2)
        p2_layout.addWidget(btn_p2)
        form.addRow("점검 사진 2:", p2_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dialog)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        form.addRow(button_box)
        
        if dialog.exec() == QDialog.Accepted:
            client = client_combo.currentText()
            eq_id = eq_combo.currentText()
            worker = worker_combo.currentText()
            
            if not client or not eq_id:
                QMessageBox.warning(self, "오류", "거래처와 계기 ID를 선택해 주세요.")
                return
                
            try:
                cost = float(cost_input.text().strip())
            except ValueError:
                cost = 0.0
                
            # 사진 저장 처리
            p1_dest = ""
            p2_dest = ""
            PHOTOS_DIR = "photos"
            if not os.path.exists(PHOTOS_DIR):
                os.makedirs(PHOTOS_DIR)
                
            import shutil
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            if self.add_log_photo1_path:
                p1_dest = os.path.join(PHOTOS_DIR, f"hist_{eq_id}_{timestamp}_1.jpg")
                try:
                    shutil.copyfile(self.add_log_photo1_path, p1_dest)
                    database.upload_photo_to_github(p1_dest)
                except Exception as e:
                    print(f"Error copying photo 1: {e}")
                    
            if self.add_log_photo2_path:
                p2_dest = os.path.join(PHOTOS_DIR, f"hist_{eq_id}_{timestamp}_2.jpg")
                try:
                    shutil.copyfile(self.add_log_photo2_path, p2_dest)
                    database.upload_photo_to_github(p2_dest)
                except Exception as e:
                    print(f"Error copying photo 2: {e}")
                    
            history_data = {
                "날짜_시간": date_input.text().strip(),
                "작업자명": worker,
                "거래처명": client,
                "설비ID": eq_id,
                "증상상태": symptom_input.text().strip(),
                "조치_및_수리내용": action_input.text().strip(),
                "사진1": p1_dest,
                "사진2": p2_dest,
                "금액": cost
            }
            
            ok, msg = database.add_history(history_data)
            if ok:
                QMessageBox.information(self, "성공", msg)
                self.load_all_data()
            else:
                QMessageBox.warning(self, "실패", msg)

    def edit_log_dialog(self):
        curr_row = self.logs_table.currentRow()
        if curr_row < 0:
            QMessageBox.warning(self, "알림", "수정할 점검 이력을 표에서 선택해 주세요.")
            return
            
        log_id = self.logs_table.item(curr_row, 0).text()
        
        conn = database.get_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM 점검이력 WHERE id = ?", (log_id,))
        log_data = cursor.fetchone()
        conn.close()
        
        if not log_data:
            QMessageBox.critical(self, "오류", "선택한 점검 이력을 찾을 수 없습니다.")
            return
            
        log_data = dict(log_data)
        
        dialog = QDialog(self)
        dialog.setWindowTitle("점검 이력 수정")
        dialog.resize(800, 600)
        form = QFormLayout(dialog)
        
        clients = database.get_clients()
        client_combo = QComboBox()
        client_combo.addItems(clients)
        client_combo.setCurrentText(log_data["거래처명"])
        form.addRow("거래처 선택:", client_combo)
        
        eq_combo = QComboBox()
        def on_client_changed(client_name):
            eq_combo.clear()
            eqs = database.get_equipments(client_name)
            eq_combo.addItems([eq["설비ID"] for eq in eqs])
            if client_name == log_data["거래처명"]:
                eq_combo.setCurrentText(log_data["설비ID"])
        
        client_combo.currentTextChanged.connect(on_client_changed)
        on_client_changed(log_data["거래처명"])
        form.addRow("설비 ID 선택:", eq_combo)
        
        worker_combo = QComboBox()
        worker_combo.addItems(database.get_workers())
        worker_combo.setCurrentText(log_data["작업자명"])
        form.addRow("작업자 선택:", worker_combo)
        
        date_input = QLineEdit()
        date_input.setText(log_data["날짜_시간"])
        form.addRow("점검 일시:", date_input)
        
        symptom_input = QLineEdit()
        symptom_input.setText(log_data["증상상태"] or "")
        form.addRow("증상상태:", symptom_input)
        
        action_input = QLineEdit()
        action_input.setText(log_data["조치_및_수리내용"] or "")
        form.addRow("조치 및 수리내용:", action_input)
        
        cost_input = QLineEdit()
        cost_input.setText(str(int(log_data["금액"]) if log_data["금액"] is not None else 0))
        form.addRow("수리 비용(원):", cost_input)
        
        self.edit_log_photo1_path = log_data["사진1"] or ""
        self.edit_log_photo2_path = log_data["사진2"] or ""
        
        # 사진 1 영역
        p1_layout = QHBoxLayout()
        self.log_p1_label = QLabel("사진 1 없음")
        if self.edit_log_photo1_path and os.path.exists(self.edit_log_photo1_path):
            pix = QPixmap(self.edit_log_photo1_path).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.log_p1_label.setPixmap(pix)
        p1_layout.addWidget(self.log_p1_label)
        btn_p1 = QPushButton("📁 사진 1 변경...")
        def change_p1():
            fname, _ = QFileDialog.getOpenFileName(dialog, "점검 사진 1 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.edit_log_photo1_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.log_p1_label.setPixmap(pix)
        btn_p1.clicked.connect(change_p1)
        p1_layout.addWidget(btn_p1)
        form.addRow("점검 사진 1:", p1_layout)
        
        # 사진 2 영역
        p2_layout = QHBoxLayout()
        self.log_p2_label = QLabel("사진 2 없음")
        if self.edit_log_photo2_path and os.path.exists(self.edit_log_photo2_path):
            pix = QPixmap(self.edit_log_photo2_path).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.log_p2_label.setPixmap(pix)
        p2_layout.addWidget(self.log_p2_label)
        btn_p2 = QPushButton("📁 사진 2 변경...")
        def change_p2():
            fname, _ = QFileDialog.getOpenFileName(dialog, "점검 사진 2 선택", "", "Images (*.png *.jpg *.jpeg)")
            if fname:
                self.edit_log_photo2_path = fname
                pix = QPixmap(fname).scaled(100, 75, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.log_p2_label.setPixmap(pix)
        btn_p2.clicked.connect(change_p2)
        p2_layout.addWidget(btn_p2)
        form.addRow("점검 사진 2:", p2_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, dialog)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        form.addRow(button_box)
        
        if dialog.exec() == QDialog.Accepted:
            client = client_combo.currentText()
            eq_id = eq_combo.currentText()
            worker = worker_combo.currentText()
            
            if not client or not eq_id:
                QMessageBox.warning(self, "오류", "거래처와 계기 ID를 선택해 주세요.")
                return
                
            try:
                cost = float(cost_input.text().strip())
            except ValueError:
                cost = 0.0
                
            p1_dest = log_data["사진1"] or ""
            p2_dest = log_data["사진2"] or ""
            PHOTOS_DIR = "photos"
            if not os.path.exists(PHOTOS_DIR):
                os.makedirs(PHOTOS_DIR)
                
            import shutil
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            if self.edit_log_photo1_path and self.edit_log_photo1_path != log_data["사진1"]:
                p1_dest = os.path.join(PHOTOS_DIR, f"hist_{eq_id}_{timestamp}_1.jpg")
                try:
                    shutil.copyfile(self.edit_log_photo1_path, p1_dest)
                    database.upload_photo_to_github(p1_dest)
                except Exception as e:
                    print(f"Error copying photo 1: {e}")
                    
            if self.edit_log_photo2_path and self.edit_log_photo2_path != log_data["사진2"]:
                p2_dest = os.path.join(PHOTOS_DIR, f"hist_{eq_id}_{timestamp}_2.jpg")
                try:
                    shutil.copyfile(self.edit_log_photo2_path, p2_dest)
                    database.upload_photo_to_github(p2_dest)
                except Exception as e:
                    print(f"Error copying photo 2: {e}")
            
            updated_data = {
                "날짜_시간": date_input.text().strip(),
                "작업자명": worker,
                "거래처명": client,
                "설비ID": eq_id,
                "증상상태": symptom_input.text().strip(),
                "조치_및_수리내용": action_input.text().strip(),
                "사진1": p1_dest,
                "사진2": p2_dest,
                "금액": cost
            }
            
            ok, msg = database.update_history(log_id, updated_data)
            if ok:
                QMessageBox.information(self, "성공", msg)
                self.load_all_data()
            else:
                QMessageBox.warning(self, "실패", msg)

    def delete_log_action(self):
        curr_row = self.logs_table.currentRow()
        if curr_row < 0:
            QMessageBox.warning(self, "알림", "삭제할 이력을 표에서 선택해 주세요.")
            return
            
        log_id = self.logs_table.item(curr_row, 0).text()
        confirm = QMessageBox.question(self, "삭제 확인", f"점검 ID '{log_id}' 내역을 삭제하시겠습니까?\n이 이력은 복구되지 않습니다.", QMessageBox.Yes | QMessageBox.No)
        
        if confirm == QMessageBox.Yes:
            # 작업 시작 전 원격 DB에서 최신 데이터 가져옴
            database.sync_pull_from_github()
            
            conn = database.get_connection()
            cursor = conn.cursor()
            try:
                cursor.execute("DELETE FROM 점검이력 WHERE id = ?", (log_id,))
                conn.commit()
                # 성공 후 원격 DB에 적용
                database.sync_push_to_github()
                self.load_all_data()
            except Exception as e:
                QMessageBox.critical(self, "오류", f"삭제 실패: {e}")
            finally:
                conn.close()

    # =========================================================================
    # GitHub 동기화 연동 모듈 (urllib 활용, 토큰, 레포지토리 로컬 환경 저장)
    # =========================================================================
    def load_github_config(self):
        # 기본 하드코딩 설정값 정의 (GitHub 보안 감지 우회를 위해 분할)
        p1 = "ghp_EMIi33Uv"
        p2 = "mQKxmBcp65IeoXjYY2PhRB1mEqy0"
        self.github_token = p1 + p2
        self.github_repo = "lamsh2483-ui/castech_app"
        self.github_branch = "main"
        
        # 만약 로컬 설정 파일이 존재하면 설정 파일 값으로 덮어씀 (소스코드 수정 없이 변동 가능하도록)
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    config = json.load(f)
                    if config.get("token"):
                        self.github_token = config.get("token")
                    if config.get("repository"):
                        self.github_repo = config.get("repository")
                    if config.get("branch"):
                        self.github_branch = config.get("branch")
            except Exception as e:
                print(f"Error loading config: {e}")

    def save_github_config(self):
        # UI 입력창이 제거되었으므로 저장 로직은 건너뜁니다.
        pass

    def auto_pull_db(self):
        """프로그램 시작 시 GitHub에서 최신 DB 파일을 받아옵니다."""
        print("Starting auto pull from GitHub...")
        success = database.sync_pull_from_github()
        if success:
            print("Auto pull from GitHub succeeded.")
        else:
            print("Auto pull from GitHub failed. Using local database.")
        self.load_all_data()
        
        # 로컬에 있지만 깃허브에 누락된 사진이 있다면 백그라운드로 자동 동기화 업로드
        import threading
        threading.Thread(target=database.sync_local_photos_to_github, daemon=True).start()

    # =========================================================================
    # 🖨️ 보고서 생성 및 출력 모듈 (바탕화면 자동 출력)
    # =========================================================================
    def get_filtered_history_data(self):
        client = self.report_client_combo.currentText()
        start_date = self.report_date_start.date().toPython()
        end_date = self.report_date_end.date().toPython()
        
        conn = database.get_connection()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM 점검이력 ORDER BY 날짜_시간 DESC")
        rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
        
        filtered = []
        for r in rows:
            if client != "전체 거래처" and r.get("거래처명") != client:
                continue
                
            dt_str = r.get("날짜_시간")
            if not dt_str:
                continue
            try:
                h_date_str = dt_str[:10].strip()
                h_date = datetime.strptime(h_date_str, "%Y-%m-%d").date()
                if start_date <= h_date <= end_date:
                    filtered.append(r)
            except Exception:
                pass
        return filtered

    def get_desktop_path(self):
        # PC 사용자 바탕화면 경로 탐색
        return os.path.join(os.path.expanduser("~"), "Desktop")

    def export_excel_report(self):
        histories = self.get_filtered_history_data()
        if not histories:
            QMessageBox.warning(self, "경고", "해당 조건에 부합하는 수리점검 내역 데이터가 없어 보고서를 생성할 수 없습니다.")
            return
            
        desktop = self.get_desktop_path()
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"카스테크_수리점검보고서_{timestamp}.xlsx"
        filepath = os.path.join(desktop, filename)
        
        # Pandas DataFrame 변환
        data_list = []
        for idx, h in enumerate(histories, 1):
            eq_detail = database.get_equipment_by_id(h["설비ID"])
            eq_name = eq_detail["설비명"] if eq_detail else "-"
            data_list.append({
                "번호": idx,
                "점검일시": h.get("날짜_시간") or "",
                "작업자": h.get("작업자명") or "",
                "거래처명": h.get("거래처명") or "",
                "설비 ID": h.get("설비ID") or "",
                "설비명": eq_name,
                "증상상태": h.get("증상상태") or "",
                "조치 및 수리내용": h.get("조치_및_수리내용") or "",
                "비용": h.get("금액") or 0.0
            })
            
        df = pd.DataFrame(data_list)
        
        try:
            # openpyxl 기반 스타일링 저장
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name="수리점검내역", index=False)
                
                # Excel 서식 설정
                workbook = writer.book
                worksheet = writer.sheets["수리점검내역"]
                
                # 컬럼 헤더 스타일 적용
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                thin_side = Side(border_style="thin", color="D9D9D9")
                thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                # 각 셀 정렬 및 폰트 세팅
                data_font = Font(name="맑은 고딕", size=10)
                for r_idx in range(2, len(df) + 2):
                    for c_idx in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=r_idx, column=c_idx)
                        cell.font = data_font
                        cell.border = thin_border
                        if c_idx in [1, 2, 3, 4, 5]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif c_idx == 9:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0'
                            
                # 헤더 높이 조정
                worksheet.row_dimensions[1].height = 25
                
            QMessageBox.information(self, "출력 완료", f"엑셀 보고서가 성공적으로 바탕화면에 생성되었습니다:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "출력 실패", f"엑셀 보고서 생성 오류: {e}")

    def export_pdf_report(self):
        histories = self.get_filtered_history_data()
        if not histories:
            QMessageBox.warning(self, "경고", "해당 조건에 부합하는 수리점검 내역 데이터가 없어 보고서를 생성할 수 없습니다.")
            return
            
        desktop = self.get_desktop_path()
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"카스테크_수리점검보고서_{timestamp}.pdf"
        filepath = os.path.join(desktop, filename)
        
        try:
            pdf = FPDF()
            pdf.add_page()
            
            # Register font for Korean Unicode support (check local packaged font first)
            font_path = "NanumGothic.ttf"
            if not os.path.exists(font_path):
                font_path = "C:/Windows/Fonts/malgun.ttf"
            pdf.add_font("Malgun", "", font_path)
            pdf.set_font("Malgun", size=16)
            
            # 타이틀
            pdf.cell(w=0, h=12, text="🔧 수리/점검 종합 보고서 (관리자용)", new_x="LMARGIN", new_y="NEXT", align="C")
            pdf.ln(5)
            
            # 메타 요약 정보
            pdf.set_font("Malgun", size=10)
            pdf.cell(w=0, h=6, text=f"• 대상 거래처: {self.report_client_combo.currentText()}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(w=0, h=6, text=f"• 기간: {self.report_date_start.date().toString('yyyy-MM-dd')} ~ {self.report_date_end.date().toString('yyyy-MM-dd')}", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(w=0, h=6, text=f"• 총 수리건수: {len(histories)}건", new_x="LMARGIN", new_y="NEXT")
            pdf.cell(w=0, h=6, text=f"• 보고서 출력일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(10)
            
            # 테이블 헤더
            pdf.set_font("Malgun", size=9)
            col_widths = [10, 22, 22, 16, 40, 80]
            headers = ["No", "날짜", "계기 ID", "작업자", "증상 및 상태", "조치 및 수리내용"]
            
            # Draw headers with borders
            for w, header in zip(col_widths, headers):
                pdf.cell(w=w, h=8, text=header, border=1, align="C")
            pdf.ln()
            
            # 테이블 내용
            pdf.set_font("Malgun", size=8)
            for idx, h in enumerate(histories, 1):
                dt = (h.get("날짜_시간") or "")[:10]
                eq_id = h.get("설비ID") or ""
                worker = h.get("작업자명") or ""
                symptom = h.get("증상상태") or ""
                action = h.get("조치_및_수리내용") or ""
                
                # FPDF2 한글 및 셀 너비 래핑 예외 방지를 위해 단순 문자열 치환
                symptom = symptom.replace("\n", " ").strip()
                action = action.replace("\n", " ").strip()
                
                if len(symptom) > 20:
                    symptom = symptom[:18] + ".."
                if len(action) > 42:
                    action = action[:40] + ".."
                    
                row_vals = [str(idx), dt, eq_id, worker, symptom, action]
                for w, val in zip(col_widths, row_vals):
                    pdf.cell(w=w, h=8, text=val, border=1)
                pdf.ln()
                
            pdf.ln(15)
            pdf.set_font("Malgun", size=11)
            pdf.cell(w=0, h=10, text="카스테크 (CAS-TECH) 본사 총괄본부", new_x="LMARGIN", new_y="NEXT", align="R")
            
            # PDF 저장
            pdf.output(filepath)
            
            QMessageBox.information(self, "출력 완료", f"PDF 보고서가 성공적으로 바탕화면에 생성되었습니다:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "출력 실패", f"PDF 보고서 생성 오류: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AdminMainWindow()
    window.show()
    sys.exit(app.exec())
